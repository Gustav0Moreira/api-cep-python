#Requesições para API
import openpyxl.workbook
import requests
import json

#Extração e formatação de dados
import openpyxl

#/-/-/-/-Funções-/-/-/-/
def Get_Cep(cep):
    #Utiliza do pacote requests para realizar um requesição na API AwesomeAPI CEP
    #O CEP será devolvido como um Dicionário contendo key:
    #'cep', 'endereco', 'bairro', 'cidade' e 'estado'

    try:
        cep_requesicao = (f'https://cep.awesomeapi.com.br/json/{cep}')
    except Exception as err:
        print(f'Ação inválida:\n{err}')
    else:
        cep_request = requests.get(cep_requesicao)
        status_code = cep_request.status_code
        if status_code == 200:

            js_cep = cep_request.json()
            info_cep = {
                'cep' : js_cep['cep'], 
                'endereco' : js_cep['address'], 
                'bairro' : js_cep['district'], 
                'cidade' : js_cep['city'], 
                'estado' : js_cep['state']
                }
            
            print(f'-/-/-/-Código CEP ({cep}) localizado com sucesso-/-/-/-')
            return info_cep

        elif status_code == 400:

            info_cep = {
                'cep' : 'Inválido', 
                'endereco' : 'Inválido', 
                'bairro' : 'Inválido', 
                'cidade' : 'Inválido', 
                'estado' : 'Inválido'
                }
            print(f'-/-/-/-Código CEP ({cep}) inválido-/-/-/-')
            return info_cep

        elif status_code == 404:

            info_cep = {
                'cep' : 'Não Localizado', 
                'endereco' : 'Não Localizado', 
                'bairro' : 'Não Localizado', 
                'cidade' : 'Não Localizado', 
                'estado' : 'Não Localizado'
                }
            print(f'-/-/-/-Código CEP ({cep}) não localizado-/-/-/-')
            return info_cep

        else:
            print("-/-/-/-Algo deu errado-/-/-/-")
            return None

def Push_Doc(dicionario):
    #Utilização do módulo openpyxl para criar e editar arquivos xlsx (Excel)
    #A função recebe o dicionário json do requesição da API CEP
    #Logo após, é criado com Workbook() um documento xlsx
    #Depois é criado a folha do documento, com o nome 'Lista_CEP'
    #Em seguida, o programa cria as células base, utilizando do append, e realiza um iteração com o for
    #para atribuir as células com os dados adquiridos da requesição da API
    #Por último, é realizado um try para salvar o arquivo e cria-lo
    cep_info = dicionario
    book = openpyxl.Workbook()
    book.create_sheet("Lista_CEP")
    selec_pag = book["Lista_CEP"]
    selec_pag.append(["CEP", "ENDEREÇO", "BAIRRO", "CIDADE", "ESTADO"])
    for i in cep_info:
        selec_pag.append([i['cep'], i['endereco'], i['bairro'], i['cidade'], i['estado']])
    try:
        book.save("documento_ceps.xlsx")
    except PermissionError:
        print("\n-/-/-/-Falha ao Criar/Salvar o documento: Documento aberto-/-/-/-\n")
    else:
        print("\n-/-/-/-Documento Criado/Atualizado com sucesso-/-/-/-\n")

#Input dos CEPs
#CEPs são retornados como list para futura iteração com map
#Casting da lista como set para evitar duplicidade
#Atribuição do map para iterar sobre a lista com a função de Get_Cep para requesição a API
#Depois da ação do map, é necessário o Casting em list para visualizar o retorno do map
while True:
    lista_cep = input('Digite os CEPs que deseja localizar:\nOBS: Os ceps devem ser digitados como este exemplo:\n"00000000 11111111 22222222 33333333"\ntodos separados por um espaço em branco.\nCEP: ').split()
    lista_cep_tratada = set(lista_cep) 
    extrato_dados_cep = map(Get_Cep, lista_cep_tratada)
    Push_Doc(list(extrato_dados_cep))

    while True:
        try:
            control = int(input("Deseja rodar o programa novamente:\n1 - Sim\n2 - Não\nRep: "))
        except ValueError:
            print("Digite uma opção válida...")
        except Exception:
            print("Algo deu errado...")
        else:
            if control == 1:
                print("O opção selecionada: 1 - Sim.")
                break
            elif control == 2:
                print("O opção selecionada: 2 - Não")

                break
            else:
                print('Digite um opção válida...')
    if control == 2:
        break
print("Até breve!")